import openpyxl
import matplotlib.pyplot as plt
import pandas as pd

wb = openpyxl.Workbook() # Call a Workbook() function of openpyxl to create a new blank Workbook object
sheet = wb.active # Get workbook active sheet from the active attribute

#This is the Main Program
expression="(x1**3-x2)**2+2*(x2-x1)**4" # The input expression is to be entered here
expression_differentials={"x1":"(6*x1*x1*(x1**3-x2)-8*(x2-x1)**3)","x2":"(-2*(x1**3-x2)+8*(x2-x1)**3)"}
hessian=["30*x1**4+24*x1**2+24*x2**2-60*x1*x2","-30*x1**2-24*x2**2+48*x1*x2","-30*x1**2-24*x2**2+48*x1*x2","2+24*x1**2+24*x2**2-48*x1*x2"]
x_k={"x1":2.0,"x2":0.0} #Dimensions with the starting values and entered here
number_of_dimensions=len(x_k)
stopping_criteria_individual_dimension={"x1":0.01,"x2":0.01} #To compare with each dimension and then decide whether to stop
stopping_criteria_functional_value=0.001

plt.title("Newtons Method")
plt.ylabel("x2")
plt.xlabel("x1")

# Newtons_Method
def Newtons_Method(expression,expression_differentials,hessian,x_k,k,stopping_criteria_functional_value=None,stopping_criteria_individual_dimension=None):
    
    if k==1:
        cell = sheet.cell(row = k, column = 1)
        cell.value = "Iteration k"
        cell = sheet.cell(row = k, column = 2)
        cell.value = "x_k"
        cell = sheet.cell(row = k, column = 3)
        cell.value = "f(x_k)"
        cell = sheet.cell(row = k, column = 4)
        cell.value = "Del_f(x_k)"
        cell = sheet.cell(row = k, column = 5)
        cell.value = "Norm_of_Del_f(x_k)"
        cell = sheet.cell(row = k, column = 6)
        cell.value = "H_k"
        cell = sheet.cell(row = k, column = 7)
        cell.value = "H_k_inverse"
        cell = sheet.cell(row = k, column = 8)
        cell.value = "-H_k_inverse*Del_f_x_k"        
        cell = sheet.cell(row = k, column = 9)
        cell.value = "x_k+1"
        cell = sheet.cell(row = k, column = 10)
        cell.value = "Stopping Criteria based on Functional Value"
        cell = sheet.cell(row = k, column = 11)
        cell.value = "Stopping Criteria based on Individual Dimensional Values"

    k+=1
    cell = sheet.cell(row = k, column = 1)
    iteration_number=k-1
    cell.value = str(iteration_number)
    cell = sheet.cell(row = k, column = 2)
    cell.value = str(x_k)

    f_x_k=expression
    for i in x_k:
        f_x_k = f_x_k.replace(str(i),"("+str(x_k[i])+")")

    cell = sheet.cell(row = k, column = 3)
    f_x_k=eval(f_x_k)
    cell.value = str(f_x_k)

    norm=0
    value_of_expression_differential=expression_differentials.copy()
    for i in expression_differentials:
        for j in x_k:
            value_of_expression_differential[i] = value_of_expression_differential[i].replace(str(j),"("+str(x_k[j])+")")
        value_of_expression_differential[i]=eval(value_of_expression_differential[i])
        norm+=value_of_expression_differential[i]*value_of_expression_differential[i]
    cell = sheet.cell(row = k, column = 4)
    cell.value = str(value_of_expression_differential)

    norm=norm**0.5
    cell = sheet.cell(row = k, column = 5)
    cell.value = str(norm)

    values_of_hessian=hessian.copy()
    for i in range(len(hessian)):
        print(values_of_hessian[i])
        for j in x_k:
            values_of_hessian[i] = values_of_hessian[i].replace(str(j),"("+str(x_k[j])+")")
        values_of_hessian[i]=eval(values_of_hessian[i])
    cell = sheet.cell(row = k, column = 6)
    cell.value = str(values_of_hessian)

    # This inverse calculation is only for 2~Dimensional Matrices and this program is not robust to be used for more than 2 functional variables.
    hessian_determinant=values_of_hessian[0]*values_of_hessian[3]-values_of_hessian[2]*values_of_hessian[1]
    if hessian_determinant==0:
        print("Determinant is Zero. Program will show Error soon")
    hessian_inverse=[values_of_hessian[3]/hessian_determinant,-values_of_hessian[1]/hessian_determinant,-values_of_hessian[2]/hessian_determinant,values_of_hessian[0]/hessian_determinant]
    cell = sheet.cell(row = k, column = 7)
    cell.value = str(hessian_inverse)

    addition=value_of_expression_differential.copy()
    sum=-1
    for i in addition:
        addition[i]=0
        for j in value_of_expression_differential:
            sum+=1
            addition[i]-=hessian_inverse[sum]*value_of_expression_differential[j]
    cell = sheet.cell(row = k, column = 8)
    cell.value = str(addition)

    x_k_1=x_k.copy()
    for i in x_k_1:
        x_k_1[i]=x_k[i]+addition[i]
    cell = sheet.cell(row = k, column = 9)
    cell.value = str(x_k_1)

    for i in x_k:
        Point_CoOrdinates[i].append(x_k_1[i])

    # The Stopping Criterias
    if stopping_criteria_functional_value!=None:
        """f_x_k_1=expression
        for i in x_k:
            f_x_k_1 = f_x_k_1.replace(str(i),"("+str(x_k_1[i])+")")
        f_x_k_1=eval(f_x_k_1)
        Stopp_Value=abs((f_x_k_1-f_x_k)/f_x_k)"""
        Stopp_Value=norm
        cell = sheet.cell(row = k, column = 10)
        cell.value = str(Stopp_Value)
        if Stopp_Value>stopping_criteria_functional_value:
            return Newtons_Method(expression,expression_differentials,hessain,x_k_1,k=k,stopping_criteria_functional_value=stopping_criteria_functional_value)

    if stopping_criteria_individual_dimension!=None:
        for_printing_stopp=stopping_criteria_individual_dimension.copy()
        stop=1
        for i in x_k:
            for_printing_stopp[i]=abs(x_k_1[i]-x_k[i])
            if for_printing_stopp[i]>stopping_criteria_individual_dimension[i]:
                stop=0
        cell = sheet.cell(row = k, column = 11)
        cell.value = str(for_printing_stopp)        
        if stop==0:
            return Newtons_Method(expression,expression_differentials,hessian,x_k_1,k=k,stopping_criteria_individual_dimension=stopping_criteria_individual_dimension)

Point_CoOrdinates={}
for i in x_k:
    Point_CoOrdinates[i]=[x_k[i]]

Newtons_Method(expression,expression_differentials, hessian, x_k, k=1, stopping_criteria_individual_dimension=stopping_criteria_individual_dimension)
#Newtons_Method(expression, expression_differentials, hessian,x_k, k=1, stopping_criteria_functional_value=stopping_criteria_functional_value)
print(Point_CoOrdinates)
wb.save("Newtons.xlsx")
 
# creating dataframe
df = pd.DataFrame(Point_CoOrdinates)
# Plotting The graph
#plt.plot(df["x1"],df["x2"])
for i in range(len(Point_CoOrdinates["x1"])):
    v=1-0.033*i
    w=0+0.033*i
    plt.scatter(Point_CoOrdinates["x1"][i],Point_CoOrdinates["x2"][i],c=(v,0,0),edgecolors=(w,0,0))

#plt.scatter(df["x1"],df["x2"],c=(),edgecolors=(1,0.4,0.3))

#routes = [(i, j) for i in Point_CoOrdinates["x1"] for j in Point_CoOrdinates["x2"]]
#arrowprops = dict(arrowstyle='->', connectionstyle='arc3', edgecolor='blue')
arrowprops = dict(arrowstyle='->', connectionstyle='arc3', edgecolor=(0.1,0.5,1))

for i in range(1,len(Point_CoOrdinates["x1"])):
    v=1-0.033*i
    arrowprops = dict(arrowstyle='->', connectionstyle='arc3', edgecolor=(0,v,0))
    plt.annotate('', xy=[Point_CoOrdinates["x1"][i], Point_CoOrdinates["x2"][i]], xytext=[Point_CoOrdinates["x1"][i-1], Point_CoOrdinates["x2"][i-1]], arrowprops=arrowprops)
plt.savefig("{}".format("Newtons Method.png"))